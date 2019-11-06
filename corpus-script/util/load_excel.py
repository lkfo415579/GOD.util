from openpyxl import load_workbook
import sys

wb = load_workbook(sys.argv[1])
alpha = sys.argv[2]


ws = wb['Xbench QA']
src_rows = ws['%s1:%s10000' % (alpha, alpha)]

for src in src_rows:
    src = src[0]
    if src.value:
        print (src.value)

